"""
zip_extractor.py
Extracts readable text from documents inside a ZIP file.

Supported formats:
  .txt  .md  .rst  .csv  .json          → direct text read
  .xlsx .xls                             → openpyxl / xlrd (auto-installed)
  .docx .doc                             → XML extraction (no deps)
  .pdf                                   → lightweight regex extraction
  .pptx                                  → XML slide text extraction

Unsupported formats (images, binaries) are silently skipped.
"""

import zipfile
import io
import os
import re
import sys
import subprocess


SUPPORTED = {".txt", ".md", ".rst", ".csv", ".json",
             ".xlsx", ".xls", ".docx", ".doc", ".pdf", ".pptx"}


def extract_documents_from_zip(zip_bytes: bytes) -> dict:
    """
    Returns {filename: text_content} for every readable file in the ZIP.
    Files with unsupported extensions or unreadable content are skipped.
    """
    documents = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            # Skip directories and macOS/Windows metadata
            base = os.path.basename(name)
            if name.endswith("/") or base.startswith("__") or base.startswith("."):
                continue

            ext = os.path.splitext(name)[1].lower()
            if ext not in SUPPORTED:
                continue

            with zf.open(name) as f:
                raw = f.read()

            text = _extract(ext, raw, name)
            if text and text.strip():
                documents[name] = text

    return documents


# ── Dispatcher ────────────────────────────────────────────────────────────────

def _extract(ext: str, raw: bytes, name: str) -> str:
    if ext in {".txt", ".md", ".rst"}:
        return _text(raw)
    if ext == ".csv":
        return _csv(raw, name)
    if ext == ".json":
        return _json(raw, name)
    if ext in {".xlsx", ".xls"}:
        return _excel(raw, ext, name)
    if ext in {".docx", ".doc"}:
        return _docx(raw)
    if ext == ".pdf":
        return _pdf(raw)
    if ext == ".pptx":
        return _pptx(raw)
    return ""


# ── Plain text ────────────────────────────────────────────────────────────────

def _text(raw: bytes) -> str:
    try:
        return raw.decode("utf-8", errors="replace")
    except Exception:
        return ""


# ── CSV ───────────────────────────────────────────────────────────────────────

def _csv(raw: bytes, name: str) -> str:
    import csv
    try:
        content = raw.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return ""

        lines = [f"[CSV file: {os.path.basename(name)}]"]
        # Header
        if rows:
            lines.append("Columns: " + ", ".join(rows[0]))
        # First 100 rows as readable text
        for row in rows[1:101]:
            lines.append(" | ".join(str(c) for c in row))
        if len(rows) > 101:
            lines.append(f"... ({len(rows)-1} total rows, showing first 100)")
        return "\n".join(lines)
    except Exception as e:
        return f"[CSV: could not parse — {e}]"


# ── JSON ──────────────────────────────────────────────────────────────────────

def _json(raw: bytes, name: str) -> str:
    import json
    try:
        data = json.loads(raw.decode("utf-8", errors="replace"))
        pretty = json.dumps(data, indent=2, ensure_ascii=False)
        # Limit to 10,000 chars to avoid flooding the LLM context
        if len(pretty) > 10000:
            pretty = pretty[:10000] + "\n... [truncated]"
        return f"[JSON file: {os.path.basename(name)}]\n{pretty}"
    except Exception as e:
        return f"[JSON: could not parse — {e}]"


# ── Excel (.xlsx / .xls) ──────────────────────────────────────────────────────

def _excel(raw: bytes, ext: str, name: str) -> str:
    # Try openpyxl for .xlsx
    if ext == ".xlsx":
        text = _excel_openpyxl(raw, name)
        if text:
            return text

    # Try xlrd for .xls (and fallback for .xlsx)
    text = _excel_xlrd(raw, name)
    if text:
        return text

    return f"[Excel file: {os.path.basename(name)} — could not read. Try saving as .csv]"


def _excel_openpyxl(raw: bytes, name: str) -> str:
    try:
        import openpyxl
    except ImportError:
        _pip_install("openpyxl")
        try:
            import openpyxl
        except ImportError:
            return ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        parts = [f"[Excel file: {os.path.basename(name)}]"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n--- Sheet: {sheet_name} ---")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if all(v is None for v in row):
                    continue
                parts.append(" | ".join("" if v is None else str(v) for v in row))
                row_count += 1
                if row_count >= 200:
                    parts.append("... [truncated at 200 rows]")
                    break
        return "\n".join(parts)
    except Exception:
        return ""


def _excel_xlrd(raw: bytes, name: str) -> str:
    try:
        import xlrd
    except ImportError:
        _pip_install("xlrd")
        try:
            import xlrd
        except ImportError:
            return ""
    try:
        wb = xlrd.open_workbook(file_contents=raw)
        parts = [f"[Excel file: {os.path.basename(name)}]"]
        for sheet in wb.sheets():
            parts.append(f"\n--- Sheet: {sheet.name} ---")
            for rx in range(min(sheet.nrows, 200)):
                row = [str(sheet.cell(rx, cx).value) for cx in range(sheet.ncols)]
                parts.append(" | ".join(row))
            if sheet.nrows > 200:
                parts.append("... [truncated at 200 rows]")
        return "\n".join(parts)
    except Exception:
        return ""


# ── DOCX ─────────────────────────────────────────────────────────────────────

def _docx(raw: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as inner:
            if "word/document.xml" not in inner.namelist():
                return ""
            xml = inner.read("word/document.xml").decode("utf-8", errors="replace")
        text = re.sub(r"<[^>]+>", " ", xml)
        text = re.sub(r"\s+", " ", text).strip()
        return text
    except Exception:
        return ""


# ── PDF ───────────────────────────────────────────────────────────────────────

def _pdf(raw: bytes) -> str:
    try:
        content = raw.decode("latin-1", errors="replace")
        parts = re.findall(r"\(([^)]{1,300})\)", content)
        lines = []
        for p in parts:
            cleaned = p.replace("\\n", "\n").replace("\\r", "").strip()
            if len(cleaned) > 3 and cleaned.isprintable():
                lines.append(cleaned)
        return "\n".join(lines[:500])
    except Exception:
        return ""


# ── PPTX ─────────────────────────────────────────────────────────────────────

def _pptx(raw: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as inner:
            slide_files = sorted(
                n for n in inner.namelist()
                if re.match(r"ppt/slides/slide\d+\.xml", n)
            )
            parts = []
            for sf in slide_files:
                xml = inner.read(sf).decode("utf-8", errors="replace")
                texts = re.findall(r"<a:t>([^<]+)</a:t>", xml)
                if texts:
                    parts.append(" ".join(texts))
            return "\n".join(parts)
    except Exception:
        return ""


# ── pip auto-install helper ───────────────────────────────────────────────────

def _pip_install(package: str):
    try:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", package],
            capture_output=True, timeout=60
        )
    except Exception:
        pass
